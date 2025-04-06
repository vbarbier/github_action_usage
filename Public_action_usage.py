import datetime
import openpyxl
import requests
from openpyxl import Workbook
from pathlib import Path

# README BEFORE RUNNING: 
# 1. Install the required libraries
# 2. Create a personal access token on Github
# 3. Fill in the parameters below
#   3.1: excelName: name of the excel file  (without the extension)
#   3.2: excelFolder: path to the folder where the excel file will be created
#   3.3: personalAccessToken: personal access token created on Github
#   3.4: filter_archived_repo: boolean to check if we want to treat the archived repositories
#   3.5: actions_list: list of actions to search
#   3.6: worksheet_name: name of the desired worksheet in the excel file
#   3.7: calls_limit_before_rate_limit_refresh: default calls number before refreshing the rate limiting information
# 4. Run the script
# The script searches for all the matches of the given action(s) in the organization afkl-airspace, 
# for each repository it has been found in, it searches for the workflow calling it (if it is the actions folder)
# And then gets the last run status and date of the workflow (if it exists) as well as the last commit date of the repository 
# and the contributors of the repository. It saves the results in an excel file with the given parameters.

# Ways to improve the script:
# 1. Add parallelism to the script to speed up the process
# 2. Add an easy way to customize the columns in the excel file
# 3. Add a simple UI to fill in the parameters

# Parameters to change
excelName = "Github_workflow_usage" # Name of the excel file
excelFolder =  r'/Path/to/your/folder/' # Path to the folder where the excel file will be created
personalAccessToken = 'ghp_XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX' # Personal access token created on Github
filter_archived_repo = True # Boolean to check if we want to skip the archived repositories
organizations_list = ['your-org-name'] # List of organizations to search
actions_list = ["name-of-the-action"] # Actions to search
worksheet_name =  datetime.datetime.today().strftime('%Y-%m-%d') # Name of the worksheet in the excel file
calls_limit_before_rate_limit_refresh = 50 # remaining calls before refreshing rate limiting information

# Excel variables
excelPath = excelFolder+excelName+'.xlsx'

# Functional variables
workflow_runs_history_cache = {}  # Cache to store the api calls history of the workflow runs
impacted_repositories = {} # List of repositories that have been impacted by the actions
rate_limit = {} # Global object with the rate limiting response
all_custom_properties = set() # Global object with all the custom properties
remaining_calls_before_refresh = calls_limit_before_rate_limit_refresh # remaining calls before refreshing rate limiting information

# Github API variables
headers = {"Authorization": "Bearer " + personalAccessToken, "Accept" : "application/vnd.github+json"}

import time

# Function to get the rate limit of the API and update the global object
def get_rate_limit():
    global rate_limit
    url = 'https://api.github.com/rate_limit'
    response = requests.get(url, headers=headers)
    if response.ok:
        rate_limit = response.json()
    else:
        return 'error'
    
# Function to refresh the rate limiting data and reset the arbitrary limit of 50 calls
def refresh_remaining_calls():
    global remaining_calls_before_refresh
    get_rate_limit()
    remaining_calls_before_refresh = calls_limit_before_rate_limit_refresh

# Function to check the rate limit, either in the saved object or by calling the API and refreshing the object and sleeping if needed
def check_rate_limit(isCodeSearch = False):
    global rate_limit
    global remaining_calls_before_refresh
    remaining_calls_before_refresh -= 1
    if isCodeSearch:
        if rate_limit['resources']['code_search']['remaining'] < 1:
            reset_time = rate_limit['resources']['code_search']['reset']
            sleep_until_reset(reset_time)
            refresh_remaining_calls()
        rate_limit['resources']['code_search']['remaining'] -= 1
    if rate_limit['rate']['remaining'] < 5:
        reset_time = rate_limit['resources']['core']['reset']
        sleep_until_reset(reset_time)
        refresh_remaining_calls()
        rate_limit['rate']['remaining'] -= 1
    if remaining_calls_before_refresh == 0:
        refresh_remaining_calls()

# Function to sleep until the reset time of the rate limit
def sleep_until_reset(reset_time):
    sleep_time = max((reset_time - time.time() + 10), 0)   # Add extra 10 seconds buffer
    time.sleep(sleep_time)
            
# Get the run history of a workflow, with a cache to avoid multiple calls
def get_workflow_history(owner, repo, workflow_name):
    key = repo+'/'+workflow_name
    if key not in workflow_runs_history_cache:
        check_rate_limit()
        url_wf='https://api.github.com/repos/'+owner+'/'+repo+'/actions/workflows/'+workflow_name+'/runs'
        response = requests.get(url_wf, headers=headers)
        if(response.ok):
            workflow_runs_history_cache[key] = response.json()
            return response.json()
        else:
            print('error valling workflow history with '+url_wf)
    else:
        return workflow_runs_history_cache[key]

# Get the last run status and date of a workflow
def get_workflow_last_status(repo, workflow_name):
    workflow_history = get_workflow_history(repo, workflow_name)
    if workflow_history and len(workflow_history['workflow_runs']) > 0:
        return {'status' : str(workflow_history['workflow_runs'][0]['status'])+': '+str(workflow_history['workflow_runs'][0]['conclusion']), 'date': str(workflow_history['workflow_runs'][0]['run_started_at'])}
    else:
        return {'status': 'N/A', 'date': 'N/A'}

# Get the last commit date of a repository    
def get_last_commit_date(owner, repo):
        url_wf='https://api.github.com/repos/'+owner+'/'+repo+'/commits'
        check_rate_limit()
        response = requests.get(url_wf, headers=headers)
        if(response.ok):
            return response.json()[0]['commit']['author']['date']
        else:
            return 'error getting commit for '+repo

# Get the contributors of a repository
def get_contributors(owner, repo_name):
    url_wf='https://api.github.com/repos/'+owner+'/'+repo_name+'/contributors'
    check_rate_limit()
    response = requests.get(url_wf, headers=headers)
    if(response.ok):
        return response.json()
    else:
        return 'error getting contributors for '+repo_name

def get_custom_properties(owner, repo_name):
    url_wf='https://api.github.com/repos/'+owner+'/'+repo_name+'/properties/values'
    check_rate_limit()
    response = requests.get(url_wf, headers=headers)
    if(response.ok):
        return response.json()
    else:
        return 'error getting properties for '+repo_name
    
# Get all the usage of an action in the organization
def get_action_usage(action_name):
    url = "https://api.github.com/search/code"
    org_filter = "(" + " OR ".join(f"org:{org}" for org in organizations_list) + ") "
    params = {"q": "org:"+org_filter+action_name+(" -is:archived" if filter_archived_repo else "")+" language:YAML", 'per_page': 100}
    results = []
    while url: # As long as we have a next value (pagination), we continue
        check_rate_limit(True)
        response = requests.get(url, headers=headers, params=params) # We call the API
        if response.ok:
            response_items = response.json()["items"]
            if 'next' in response.links: # If there is a next value in the answer, it means we have pagination and we keep the info to iterate
                url = response.links['next']['url']
            else: # Else we put None and it stops
                url = None
            for result in response_items: # For each result, we process it
                repo_url = result["repository"]["html_url"]
                repo_name = repo_url.split("/")[-1]
                file_path = result['path']
                repo_owner = repo_url.split("github.com/")[1].split("/")[0]
                custom_properties = get_custom_properties(repo_owner, repo_name)
                for property in custom_properties: # Add the custom properties to the global set so we can have a list of all the custom properties
                    all_custom_properties.add(property['property_name'])
                if repo_name not in impacted_repositories: # Here we get the last commit date and the contributors of the repository only it it's not already in the list
                    contributors = get_contributors(repo_owner, repo_name)
                    contributors_email = ', '.join(
                        [contributor['login'].replace('_afklm', '@airfranceklm.com')
                         for contributor in contributors
                         if not contributor['login'].startswith('PAT') and contributor['login'].endswith('_afklm')]
                    )
                    impacted_repositories[repo_name] =   [get_last_commit_date(repo_owner, repo_name), contributors_email]
                print(custom_properties)
                currentResult = {'org': repo_owner, 'repository': repo_name, 'repository_url': repo_url, 'file': file_path, 'last_commit': impacted_repositories[repo_name][0], 'contributors': impacted_repositories[repo_name][1], 'custom_properties': custom_properties}
                results.append(currentResult)
        else:
            print("Error during request : %s" % response.status_code)
    return results

# Recursive function to get the path of the action in the repository, by appending the path of the file to the previous path until a workflow is found            
def get_action_path(owner, paths, action_name, repo_name, final_result_list):
    url = "https://api.github.com/search/code"    
    params = {'q': 'repo:'+owner+'/' +repo_name +' "uses: ./'+action_name+ '" language:YAML', 'per_page': 100}
    while url: # As long as we have a next value (pagination), we continue
        check_rate_limit(True)
        response = requests.get(url, headers=headers, params=params) # We call the API
        if response.ok:
            if(response.json()['total_count'] == 0): # Case where there is no result for this action
                final_result_list.append({'call_chain': paths, 'workflow': 'none'})
            for item in response.json()["items"]: # If we have results, we process them
                file_path = item['path']
                full_path = paths+', '+file_path
                if file_path.startswith('.github/actions'): # If the action is in the actions folder, we continue the search
                    get_action_path(full_path, file_path[:file_path.rfind('/')], repo_name, final_result_list)
                elif file_path.startswith('.github/workflows'): # If the action is in the workflows folder, we add the result to the final list
                    final_result_list.append({'call_chain': full_path, 'workflow': file_path.split('/')[-1]})
                else:
                    # Code for other paths, rare cases where the action is not in the .github/actions or .github/workflows
                    # It seems to happen when the action is in templates repositories
                    print('The action is not in an usual folder, maybe you should check. Repository and path: '+repo_name+' '+file_path)
            if 'next' in response.links: # If there is a next value in the answer, it means we have pagination and we keep the info to iterate
                url = response.links['next']['url']
            else: # Else we put None and it stops
                url = None
    return final_result_list

# Function to setup the excel file, by creating or getting the workbook and the worksheet
def setup_excel():
    global worksheet_name
    global excelPath
    # Check if the file already exists, create it if needed
    if not Path(excelPath).is_file():
        workbook = Workbook()
        workbook.save(excelPath)
        print(f"File {excelPath} created")
    # Load the workbook 
    workbook = openpyxl.load_workbook(excelPath)
    # Check if the worksheet already exists, and use a suffix if it does, then create it
    suffix = 0
    while worksheet_name in workbook.sheetnames:
        print(f"Worksheet {worksheet_name} already exists, adding a suffix")
        suffix += 1
        worksheet_name = f"{worksheet_name}_{suffix}"
    worksheet = workbook.create_sheet(title=worksheet_name)
    return workbook, worksheet

# Save the results in the excel file
def write_results_to_excel(workbook, worksheet, final_results):
    # Define the headers based on the keys of the first dictionary in final_results
    headers = final_results[0].keys()
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    # Add the data to the worksheet
    for row_num, result in enumerate(final_results, 2):
        for col_num, (key, value) in enumerate(result.items(), 1):
            worksheet.cell(row=row_num, column=col_num, value=value)
    # Save the workbook
    workbook.save(excelPath)

# Core function to get all the usage of the actions in the organization
def get_all_usage_for_actions_in_org():
    final_results = []
    get_rate_limit() # Initialize the rate limit in case it's not the default value
    for action in actions_list: # For each action in the list we defined
        findings = get_action_usage(action) # We get each usage of the action in the organization
        for finding in findings: # For each usage
            result_list = []
            file_path = finding['file']
            repository = finding['repository']
            owner = finding['org']
            if file_path.startswith('.github/actions'): # If the action is in the actions folder, we search for a workflow calling it
                result_list = get_action_path(owner, file_path, file_path[:file_path.rfind('/')], repository, [])
            elif file_path.startswith('.github/workflows'): # If the action is in the workflows folder, we add the result to the final list
                result_list = [{'call_chain': file_path, 'workflow': file_path.split('/')[-1]}]
            for result in result_list: # For each result, we get the last run status and date of the workflow (if it exists)
                custom_properties_dict = {property_name: '' for property_name in all_custom_properties}
                for property in finding['custom_properties']:
                    custom_properties_dict[property['property_name']] = property['value']
                workflow_status = {'status': 'N/A', 'date': 'N/A'} if result['workflow'] == 'none' else get_workflow_last_status(repository, result['workflow'])
                final_results.append({'organization': owner, 'repository': repository, 'repository_url': finding['repository_url'], 'action': action, 'call_chain': result['call_chain'], 'workflow': result['workflow'], 'workflow_last_run': workflow_status['date'], 'workflow_last_status': workflow_status['status'], 'repo_last_commit_date': finding['last_commit'], 'contributors': finding['contributors'], **custom_properties_dict})
    return final_results

final_results = get_all_usage_for_actions_in_org()
workbook, worksheet = setup_excel()
write_results_to_excel(workbook, worksheet, final_results)